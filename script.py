import zipfile
import shutil
import openpyxl
import xlrd
import os
import sys

mydir = "C://Users//Surya-Sunkari//Documents//"

#unzip file
def unzip_file(file_name, new_name):
    try:
        with zipfile.ZipFile(mydir + file_name, 'r') as zip_ref:
            zip_ref.extractall(mydir + new_name)
    except:
        print('Error when unzipping zipped file')
        sys.exit()

#delete folder
def delete_folder(folder_name):
    try:
        shutil.rmtree(mydir + folder_name)
    except OSError as e:
        print('Error: %s - %s.' % (e.filename, e.strerror))
        sys.exit()

#increase row
def incr_row(row):
    if (row+2)%5 == 0:
        return row + 3
    else:
        return row + 1
    
#get file names
main_file_name = input('Enter the name of the Excel file: ')
zipped_folder_name = input('Enter the name of the zipped folder: ')
cur_time = input('Enter the time of the first interval in 24hr time (ex "13:55:00"): ')
sheetname_dict = {'Class 1':'Medium Trucks', 'Class 2':'Cars', 'Class 3':'Large Trucks', 'Class 4':'Motorcycles', 'Class 5':'Bicycles', 'Class 6':'Truck with Trailer', 'Class 7':'Large Truck with Trailer', 'Class 8':'Bus', 'Class 9':'Tractor'}
excelfile = openpyxl.load_workbook(mydir + main_file_name + '.xlsx')
excelfile_sheetnames = excelfile.sheetnames
cur_row = 11
temp_sheet = excelfile['Medium Trucks']

try:
    while str(temp_sheet['A' + str(cur_row)].value) != cur_time:
        cur_row += 1;
except:
    print('Time given is not in the Excel file. Make sure to follow HH:MM:SS format.')
    sys.exit()

#unzip zipped folder and store in new folder
unzip_file(zipped_folder_name + '.zip', 'Unzipped_Folder')

#iterate through each file
kbinput = input('Current time: {}. Enter sheet number, "z" for zeroes, or "done" when finished: '.format(cur_time)).lower()
while kbinput != 'done':
    
    class_to_data = {'Class 1':[], 'Class 2':[], 'Class 3':[], 'Class 4':[], 'Class 5':[], 'Class 6':[], 'Class 7':[], 'Class 8':[], 'Class 9':[]}
    if kbinput == 'z':
        for key in class_to_data:
            class_to_data[key] = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    else:
        workbook = xlrd.open_workbook(f'{mydir}Unzipped_Folder//{kbinput}.xls')
        for key in class_to_data:
            cur_sheet = workbook.sheet_by_name(key)
            mylist = []            
            try:
                for col in range(1, 17):
                    mylist.append(int(cur_sheet.cell(10, col).value) + int(cur_sheet.cell(11, col).value))
            except:
                for col in range(1, 17):
                    mylist.append(int(cur_sheet.cell(10, col).value))
            
            class_to_data[key] = mylist
            
    #get data from class_to_data dict into excelfile
    for classname in class_to_data:
        sheet = excelfile[sheetname_dict[classname]]
        data = class_to_data[classname]
        for index,letter in enumerate('BCDEFGHIJKLMNOPQ'):
            sheet[letter + str(cur_row)] = data[index] 
    
    cur_row = incr_row(cur_row)
    cur_time = temp_sheet['A' + str(cur_row)].value
    
    #save the file to make sure edits are committed
    excelfile.save(main_file_name + ".xlsx")

    kbinput = input('Current time: {}. Enter sheet number, "z" for zeroes, or "done" when finished: '.format(cur_time)).lower()


#delete folder at end
delete_folder('Unzipped_Folder')
os.remove(zipped_folder_name + '.zip')